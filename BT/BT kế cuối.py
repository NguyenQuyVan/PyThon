#1 Biểu đồ cột
'''
import matplotlib.pyplot as plt

Data = { 'Cam':50, 'Xoài':75, 'Mít':500, 'Ổi':192, 'Vải':178, 'táo':300, 'Lê':200}
plt.bar( range(len(Data)), list(Data.values()), label = "số lượng", color = "r")
plt.legend()
plt.xticks( range(len(Data)), Data.keys())
plt.xlabel( "loại trái cây")
plt.ylabel( "Số lượng")
plt.title( "Biểu đồ thống kê số lượng trái cây")
plt.show()
'''
#1.2 Biểu đồ tròn
'''
import matplotlib.pyplot as plt


Data = {'Cam':50, 'Xoài':75, 'Mít':500, 'Ổi':192, 'Vải':178, 'Táo':300, 'Lê':200}

plt.pie(list(Data.values()), labels=list(Data.keys()), autopct='%1.1f%%')

plt.title("Biểu đồ thống kê số lượng trái cây")

plt.show()
'''
#Bài 1,b Tính tổng tiền các loại trái cây trên

'''
Data = [[1,'Cam', 'VN', 50, 2],
        [2,'Xoài', 'Ấn Độ', 75, 3],
        [3, 'Mít', 'VN', 500, 45],
        [4, 'Ổi', 'Nhật', 192, 7],
        [5, 'Vải', 'VN', 178, 5],
        [6, 'Táo', 'Mỹ', 300, 6],
        [7, 'Lê', 'Hàn Quốc', 200, 8]]

for i in range(7):
    print("Tổng tiền:", Data[i][1], "=", Data[i][3] * Data[i][4])
'''
#Bài 1,c
'''
#Cách 1 lưu file xls

from xlwt import Workbook



wB = Workbook()

trang_tinh_1 = wB.add_sheet('Trang tính 1')


duLieu = (

    ['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],
    
    [1, 'Cam', 'VN', 50, 2, ''],

        [2, 'Xoài', 'Ấn Độ', 75, 3, ''],
    
    [3, 'Mít', 'VN', 500, 45, ''],
    
    [4, 'Ổi', 'Nhật', 192, 7, ''],
    
    [5, 'Vải', 'VN', 178, 5, ''],
    
    [6, 'Táo', 'Mỹ', 300, 6, ''],

        [7, 'Lê', 'Hàn Quốc', 200, 8, ''],
)



row, column = 0, 0

for sTT, ten, noiTrong, soLuong, gia, tong in (duLieu):
    
    trang_tinh_1.write(row, column, sTT)
    
    trang_tinh_1.write(row, column + 1, ten)
        trang_tinh_1.write(row, column + 2, noiTrong)
        trang_tinh_1.write(row, column + 3, soLuong)
        trang_tinh_1.write(row, column + 4, gia)
        trang_tinh_1.write(row, column + 5, tong)
        
        row += 1


wB.save('D:\\Documents\\Code\\Python\\LuuExcel\\abc.xls')
'''
'''
#Cách 2 lưu file xlsx
#pip install xlsxwriter
import xlsxwriter


wB = xlsxwriter.Workbook('C:\\Users\\admin\\Desktop\\word and learn\\Python 3.9\\bài tập\\BTpython.xlsx]')

trang_tinh_1 = wB.add_worksheet('Trang tính 1')


duLieu = (

    ['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],

    [1, 'Cam', 'VN', 50, 2, ''],

    [2, 'Xoài', 'Ấn Độ', 75, 3, ''],

    [3, 'Mít', 'VN', 500, 45, ''],

    [4, 'Ổi', 'Nhật', 192, 7, ''],

    [5, 'Vải', 'VN', 178, 5, ''],

    [6, 'Táo', 'Mỹ', 300, 6, ''],

    [7, 'Lê', 'Hàn Quốc', 200, 8, ''],
)


row, column = 0, 0


for sTT, ten, noiTrong, soLuong, gia, tong in (duLieu):
    trang_tinh_1.write(row, column, sTT)

    trang_tinh_1.write(row, column + 1, ten)

    trang_tinh_1.write(row, column + 2, noiTrong)

    trang_tinh_1.write(row, column + 3, soLuong)

    trang_tinh_1.write(row, column + 4, gia)

    trang_tinh_1.write(row, column + 5, tong)

    row += 1


wB.close()
'''
# bài 1,d
'''
import pandas as pd
dfs = pd.read_excel("F:\\van\\python\\Book3.xls")
#Đọc chỉ 3 dòng
#print(dfs.head(3))
#Đọc hết
print(dfs)
'''
#Bài 1.e
'''
#pip install openpyxl
import openpyxl

def getValueExcel(fileName, cellName):
    wB = openpyxl.load_workbook(fileName)
    trang_tinh_1 = wB['Trang tính 1']
    wB.close()
    return trang_tinh_1[cellName].value

def updateValueExcel(fileName, cellName, value):
    wB = openpyxl.load_workbook(fileName)
    trang_tinh_1 = wB['Trang tính 1']
    trang_tinh_1[cellName].value = value
    wB.close()
    wB.save(fileName)

fileName = "C:\\Users\\admin\\Desktop\\word and learn\\Python 3.9\\bài tập\\BTpython.xlsx"

tongList = []

for i in range(2, 8 + 1):
    tongList.append(getValueExcel(fileName, 'D' + str(i))
                    * getValueExcel(fileName, 'E' + str(i)))
    
for i in range(len(tongList)):
    updateValueExcel(fileName, 'F' + str(i + 2), tongList[i])
'''
#Bài 1.f
import openpyxl

def getValueExcel(fileName, cellName):
    wB = openpyxl.load_workbook(fileName)
    trang_tinh_1 = wB['Trang tính 1']
    wB.close()
    return trang_tinh_1[cellName].value

fileName = "C:\\Users\\admin\\Desktop\\word and learn\\Python 3.9\\bài tập\\BTpython.xlsx"

#Tìm max
max = getValueExcel(fileName, 'F' + str(2))

for i in range(2, 8 + 1):
    if (max < getValueExcel(fileName, 'F' + str(i))):
        max = getValueExcel(fileName, 'F' + str(i))
print('Kết quả lớn nhất là: ' + str(max))

#Tìm min
min = getValueExcel(fileName, 'F' + str(2))

for i in range(2, 8 + 1):
    if (min > getValueExcel(fileName, 'F' + str(i))):
        min = getValueExcel(fileName, 'F' + str(i))
print('Kết quả nhỏ nhất là: ' + str(min))

#Tính trung bình
tongList = []
for i in range(2, 8 + 1):
    tongList.append(getValueExcel(fileName, 'F' + str(i)))

def tinhTrungBinh(numbers):
    return round(float(sum(numbers) / len(numbers)), 2)
print("Kết quả trung bình là:", tinhTrungBinh(tongList))


        
