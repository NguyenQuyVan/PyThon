#1 a bieu do hinh tron
'''
import matplotlib.pyplot as va

dat = {'Lài':100, 'Cúc':820, 'Lan':475, 'Huê':624, 'Giấy':752, 'Mai':428, 'Tulip':354}
va.pie(list(dat.values()), labels=list(dat.keys()), autopct='%0.1f%%')
va.title("Danh sách các loại hoa")
va.show()
'''

# a bieu do hinh cot

import matplotlib.pyplot as plt


'''
Data = {'Hoa Lài':100, 'Hoa Cúc':820, 'Hoa Lan':475, 'Hoa Huệ':624, 'Hoa Giấy':752, 'Hoa Mai':478, 'Hoa Tulip':354}

plt.bar(range(len(Data)), list(Data.values()), label = "SỐ LƯỢNG", color = "black")

plt.legend()

plt.xticks(range(len(Data)), Data.keys())

plt.xlabel("TÊN HOA")

plt.ylabel("SỐ LƯỢNG")

plt.title("BIỂU ĐỒ SỐ LƯỢNG HOA")

plt.show()
'''
# b Tổng giá tiền các loại hoa
'''
Data = [
    ['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],
    [1,'Hoa Lài', 'Đà Lạt', 100, 10, ''],
    [2,'Hoa Cúc', 'Cần Thơ', 820, 15, ''],
    [3, 'Hoa Lan', 'Long An', 475, 30, ''],
    [4, 'Hoa Huệ', 'Tiền Giang', 624, 45, ''],
    [5, 'Hoa Giấy', 'Bến Tre', 752, 78, ''],
    [6, 'Hoa Mai', 'Vĩnh Long', 478, 92, ''],
    [7, 'Hoa Tulip', 'Đà Lạt', 354, 47, ''],
]

for i in range(1, len(Data)):
    print("Tổng tiền:", Data[i][1], "=", Data[i][3] * Data[i][4])
'''
#c lưu bảng dữ liệu xuống 1 file tên là xyz.xlsx
'''
import xlsxwriter

wB = xlsxwriter.Workbook('C:\\Users\\Quy Van\\Desktop\\Learn\\bài tập\\xyz.xlsx')

trang_tinh_1 = wB.add_worksheet('Sheet1')

duLieu =( 
    ['STT', 'Tên', 'Nơi trồng', 'Số lượng', 'Giá', 'Tổng'],
    [1,'Hoa Lài', 'Đà Lạt', 100, 10, ''],
    [2,'Hoa Cúc', 'Cần Thơ', 820, 15, ''],
    [3, 'Hoa Lan', 'Long An', 475, 30, ''],
    [4, 'Hoa Huệ', 'Tiền Giang', 624, 45, ''],
    [5, 'Hoa Giấy', 'Bến Tre', 752, 78, ''],
    [6, 'Hoa Mai', 'Vĩnh Long', 478, 92, ''],
    [7, 'Hoa Tulip', 'Đà Lạt', 354, 47, ''],
)

row, column = 0, 0

for sTT, ten, noiTrong, soLuong, donGia, tong in (duLieu):
    trang_tinh_1.write(row, column, sTT)

    trang_tinh_1.write(row, column + 1, ten)

    trang_tinh_1.write(row, column + 2, noiTrong)

    trang_tinh_1.write(row, column + 3, soLuong)

    trang_tinh_1.write(row, column + 4, donGia)

    trang_tinh_1.write(row, column + 5, tong)

    row += 1

wB.close()
'''
#d đọc dữ liệu file xyz.xlsx và in ra màn h́nh
'''
import pandas as pd
print(pd.read_excel("C:\\Users\\Quy Van\\Desktop\\Learn\\bài tập\\xyz.xlsx"))
'''
# e cột tổng ở câu b lưu xuống file xyz
'''
import openpyxl

def getValueExcel(fileName, cellName):
    wB = openpyxl.load_workbook(fileName)
    trang_tinh_1 = wB['Sheet1']
    wB.close()
    return trang_tinh_1[cellName].value

def updateValueExcel(fileName, cellName, value):
    wB = openpyxl.load_workbook(fileName)
    trang_tinh_1 = wB['Sheet1']
    trang_tinh_1[cellName].value = value
    wB.close()
    wB.save(fileName)

fileName = "C:\\Users\\Quy Van\\Desktop\\Learn\\bài tập\\xyz.xlsx"

tongList = []

for i in range(2, 8 + 1):
    tongList.append(getValueExcel(fileName, 'D' + str(i))
                    * getValueExcel(fileName, 'E' + str(i)))
    
for i in range(len(tongList)):
    updateValueExcel(fileName, 'F' + str(i + 2), tongList[i])
'''
#F tim loại hoa đắt nhất và số lượng ít nhất

import openpyxl
'''
def getValueExcel(fileName, cellName):
    wB = openpyxl.load_workbook(fileName)
    trang_tinh_1 = wB['Sheet1']
    wB.close()
    return trang_tinh_1[cellName].value

fileName = "C:\\Users\\Quy Van\\Desktop\\Learn\\bài tập\\xyz.xlsx"
#Tìm max
max = getValueExcel(fileName, 'E' + str(2))
viTri = 0

for i in range(2, 8 + 1):
    if (max < getValueExcel(fileName, 'E' + str(i))):
        max = getValueExcel(fileName, 'E' + str(i))
        viTri = i
print('Hoa đắt nhất là: ' + getValueExcel(fileName, 'B' + str(viTri)), '-', max)
#Tìm min
min = getValueExcel(fileName, 'F' + str(2))
viTri2 = 0

for i in range(2, 8 + 1):
    if (min > getValueExcel(fileName, 'D' + str(i))):
        min = getValueExcel(fileName, 'D' + str(i))
        viTri2 = i
print('Hoa có số lượng ít nhất là: ' + getValueExcel(fileName, 'B' + str(viTri2)), '-', min)
'''
#Hãy nhập số nguyên n, tạo một list gồm các số fibonacci nhỏ hơn n và in ra
#Dãy fibonacci là dãy số nguyên được định nghĩa một cách đệ quy như sau:
#f(0)=0, f(1) = 1, f(n) = f(n-1) + f(n-2)

def f(n):
    if (n == 0):
        return 0
    elif (n == 1):
        return 1
    else:
        return f(n - 1) + f(n - 2)
       
n = int(input("Hay nhap n: "))

listN = []

for i in range(n):
    if (f(i) < n):
        listN.append(f(i))
               
print(listN)





