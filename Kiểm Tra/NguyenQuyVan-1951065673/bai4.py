#Biểu đồ hình cột
import matplotlib.pyplot as plt
data={'Điện thoại':10, 'Máy Tính':15, 'Ipad':48,'Tivi':78, 'Bếp điện':92, 'Bàn ủi':145, 'Tủ Lạnh':178,}
plt.bar(range(len(data)),list(data.values()))
plt.xticks(range(len(data)),data.keys())
plt.title("Biểu đồ cột thống kê số lượng thiết bị")
plt.show();

#Biểu đồ tròn
import matplotlib.pyplot as mp
data={'Điện thoại':10,'Máy Tính':15,'Ipad':48,'Tivi':78,'Bếp điện':92,'Bàn ủi':145,'Tủ Lạnh':178};
mp.pie(list(data.values()), labels=list(data.keys()),autopct='%1.1f%%');
mp.title("Biểu đồ thống kê số lượng sản phẩm");
mp.show();

#Tính tổng các loại hoa
data=[[1,'Điện thoại','Việt Nam',120,10],
      [2,'Máy Tính','Việt Nam',400,15],
      [3,'Ipad','Thái Lan',260,48],
      [4,'Tivi','Trung Quốc',150,78],
      [5,'Bếp điện','Trung Quốc',99,92],
      [6,'Bàn ủi ','Thái Lan',124,145],
      [7,'Tủ lạnh','Thái Lan',580,178]];
for i in range(7):
    print("Tổng tiền:",data[i][1],"=",data[i][3]*data[i][4]);

#c
import pandas as pd
print(pd.read_excek("D:\\abc.xlsx"))
import xlsxwriter

wB=xlsxwriter.Workbook('D:\\abc.xlsx')
trang_tinh_1=wB.add_worksheet('Sheet1')

dulieu=(
    ['STT','Tên','Nơi trồng','Số lượng','Giá','Tổng'],
    [1,'Điện thoại','Việt Nam',120,10,""],
    [2,'Máy Tính','Việt Nam',400,15,""],
    [3,'Ipad','Thái Lan',260,48,""],
    [4,'Tivi','Trung Quốc',150,78,""],
    [5,'Bếp điện','Trung Quốc',99,92,""],
    [6,'Bàn ủi ','Thái Lan',124,145,""],
    [7,'Tủ lạnh','Thái Lan',580,178,""],
)
row,column=0,0

for sTT,ten,noiTrong,soLuong,donGia,tong in (dulieu):
     trang_tinh_1.write(row,column,sTT)
     trang_tinh_1.write(row,column + 1,ten)
     trang_tinh_1.write(row,column + 2,noiTrong)
     trang_tinh_1.write(row,column + 3,soLuong)
     trang_tinh_1.write(row,column + 4,donGia)
     trang_tinh_1.write(row,column + 5,tong)
     row+=1
wB.close();
#d
import openpyxl
def getValueExcel(fileName, cellName):
    wB=openpyxl.load_workbook(fileName)
    trang_tinh_1=wB['Sheet1']
    wB.close()
    return trang_tinh_1[cellName].value
def updateValueExcel(fileName, cellName, value):
    wB=openpyxl.load_workbook(fileName)
    trang_tinh_1=wB['Sheet1']
    trang_tinh_1[cellName].value = value
    wB.close()
    wB.save(fileName)
fileName="D:\\abc.xlsx"
tongList=[]
for i in range(2,8+1):
    tongList.append(getValueExcel(fileName,'D'+str(i))
                    *getValueExcel(fileName,'E'+str(i)))
for i in range(len(tongList)):
    updateValueExcel(fileName,'F'+str(i+2),tongList[i])
    

